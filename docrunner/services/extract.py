"""
docrunner/services/extract.py

Extracts the placeholder schema from an uploaded template. Called by the upload
view; the result is stored on Template.schema and drives forms, validation, and
the API contract.

Schema shape (data-model §3.1):
    {
      "fields":    [{"name": "client", "type": "string", "required": true}, ...],
      "loops":     [{"name": "items"}],
      "conds":     ["paid"],
      "has_loops": true
    }

DOCX uses docxtpl's Jinja AST introspection (handles run-splitting correctly).
XLSX is regex-scanned cell by cell.
"""
import re

from docrunner.models import TemplateFormat

_PLACEHOLDER = re.compile(r"\{\{\s*([\w.]+)\s*\}\}")
_LOOP = re.compile(r"\{%\s*for\s+\w+\s+in\s+(\w+)")
_COND = re.compile(r"\{%\s*if\s+([\w.]+)")

# Heuristic type inference from a placeholder's name.
_NUMBER_HINTS = ("amount", "total", "qty", "quantity", "price", "count", "number", "num")
_DATE_HINTS = ("date", "_at", "due", "deadline", "day")
_BOOL_HINTS = ("is_", "has_", "paid", "active", "enabled")


def extract_schema(path: str, fmt: str) -> dict:
    """Dispatch on format and return the normalized schema dict."""
    if fmt == TemplateFormat.DOCX:
        names, loops, conds = _extract_docx(path)
    elif fmt == TemplateFormat.XLSX:
        names, loops, conds = _extract_xlsx(path)
    else:
        raise ValueError(f"unsupported template format: {fmt}")

    # A field that only appears as `item.x` inside a loop is not a top-level field.
    loop_vars = set(loops)
    top_level = sorted(
        n for n in names
        if "." not in n and n not in loop_vars and n not in conds
    )

    fields = [
        {"name": n, "type": _infer_type(n), "required": True}
        for n in top_level
    ]
    return {
        "fields": fields,
        "loops": [{"name": l} for l in sorted(loop_vars)],
        "conds": sorted(set(conds)),
        "has_loops": bool(loop_vars),
    }


def _extract_docx(path: str) -> tuple[set, list, list]:
    """Use docxtpl to get declared variables; scan the document's XML for the
    loop/conditional structure (`{% for %}`, `{% if %}`)."""
    from docxtpl import DocxTemplate

    doc = DocxTemplate(path)
    # get_undeclared_template_variables() resolves vars even when Word splits a
    # placeholder across runs — the whole reason we use docxtpl over raw replace.
    names = set(doc.get_undeclared_template_variables())

    # Scan the underlying document part's XML for structural tags. docxtpl's
    # get_xml() isn't reliable pre-render, so read the docx body XML directly.
    xml = _docx_body_xml(path)
    loops = _LOOP.findall(xml)
    conds = _COND.findall(xml)
    return names, loops, conds


def _docx_body_xml(path: str) -> str:
    """Return the concatenated text of word/document.xml (and headers/footers),
    with XML tags stripped so split Jinja tags rejoin for regex scanning."""
    import zipfile

    parts = []
    with zipfile.ZipFile(path) as zf:
        for entry in zf.namelist():
            if entry.startswith("word/") and entry.endswith(".xml"):
                raw = zf.read(entry).decode("utf-8", errors="replace")
                # Strip XML tags so a `{% for %}` split across <w:r> runs becomes
                # contiguous text the LOOP/COND regexes can match.
                parts.append(re.sub(r"<[^>]+>", "", raw))
    return "\n".join(parts)


def _extract_xlsx(path: str) -> tuple[set, list, list]:
    """Regex-scan every string cell across all worksheets."""
    from openpyxl import load_workbook

    names: set[str] = set()
    loops: list[str] = []
    conds: list[str] = []
    wb = load_workbook(path, read_only=True, data_only=False)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str) and "{{" in value:
                    names.update(_PLACEHOLDER.findall(value))
                    loops.extend(_LOOP.findall(value))
                    conds.extend(_COND.findall(value))
    wb.close()
    return names, loops, conds


def _infer_type(name: str) -> str:
    """Best-effort type guess from the placeholder name. Defaults to string."""
    low = name.lower()
    if any(h in low for h in _BOOL_HINTS):
        return "boolean"
    if any(low.endswith(h) or h in low for h in _DATE_HINTS):
        return "date"
    if any(h in low for h in _NUMBER_HINTS):
        return "number"
    return "string"
