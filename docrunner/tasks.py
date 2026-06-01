"""
docrunner/tasks.py

The generation worker. Every trigger (form, API, schedule, webhook) enqueues
`run_generation(run_id)` on django-q2; nothing generates inline in a request.

Flow (mirrors DocRunner-spec.md §6 and data-model §5):
    QUEUED -> RUNNING -> fill -> [convert] -> store outputs -> finalize
            -> SUCCESS | PARTIAL | FAILED -> enqueue deliveries

This module owns ALL Run status transitions. Views/endpoints only create the
Run as QUEUED and enqueue the task.
"""
import io
import os
import re
import time
import zipfile
import tempfile
import subprocess
from datetime import datetime, timezone

from django.core.files.base import ContentFile
from django.db import transaction
from django_q.tasks import async_task

from docrunner.models import (
    Run, Output, Delivery, Template,
    RunStatus, OutputFormat, TemplateFormat, MissingPolicy, DeliveryStatus,
)

# Intermediate office format we must produce before any PDF conversion.
_INTERMEDIATE = {
    TemplateFormat.DOCX: "docx",
    TemplateFormat.XLSX: "xlsx",
}

_FILENAME_SANITIZE = re.compile(r"[^A-Za-z0-9._-]+")


# ──────────────────────────────────────────────
# Entry point (enqueued with the run's UUID)
# ──────────────────────────────────────────────
def run_generation(run_id: str) -> None:
    """Execute one Run end to end. Safe to retry: re-running a finished Run is a no-op."""
    run = Run.objects.select_related("template", "binding").get(id=run_id)

    if run.status not in (RunStatus.QUEUED,):
        # Already picked up / finished. Avoid double-processing on a redelivered task.
        _log(run, f"[WARN] run_generation called on status={run.status}; skipping")
        return

    started = time.monotonic()
    run.status = RunStatus.RUNNING
    run.started_at = _now()
    run.report = {"missing": {}, "extra": {}, "errors": [], "log": []}
    run.save(update_fields=["status", "started_at", "report"])
    _log(run, f"[INFO] Starting run {run.id} ({run.template.name}, →{run.output_format})")

    try:
        outputs = _generate(run)
    except Exception as exc:  # hard failure: nothing usable produced
        _finalize(run, RunStatus.FAILED, started)
        _log(run, f"[ERROR] {type(exc).__name__}: {exc}")
        run.save(update_fields=["report"])
        raise  # let django-q2 record the failure too

    # Decide success vs partial from per-record errors.
    had_errors = bool(run.report["errors"])
    produced = bool(outputs)
    if produced and had_errors:
        status = RunStatus.PARTIAL
    elif produced:
        status = RunStatus.SUCCESS
    else:
        status = RunStatus.FAILED

    _finalize(run, status, started)
    _log(run, f"[OK]   Done: {len(outputs)} file(s), status={status} "
              f"in {run.duration_ms} ms")
    run.save(update_fields=["report"])

    if status in (RunStatus.SUCCESS, RunStatus.PARTIAL):
        _enqueue_deliveries(run, outputs)


# ──────────────────────────────────────────────
# Core generation
# ──────────────────────────────────────────────
def _generate(run: Run) -> list[Output]:
    """Fill the template once per record, convert if needed, persist Output rows.
    Returns the created Output rows (excludes the archive)."""
    template = run.template
    records = run.input_records or []
    run.record_count = len(records)
    run.save(update_fields=["record_count"])

    required = template.placeholder_names
    policy = run.binding.missing_policy if run.binding else MissingPolicy.ERROR
    filename_field = (run.binding.filename_field if run.binding else "") or ""

    _log(run, f"[INFO] {len(records)} record(s); {len(required)} placeholder(s)")

    outputs: list[Output] = []
    with tempfile.TemporaryDirectory() as workdir:
        for idx, record in enumerate(records):
            try:
                _validate_record(run, idx, record, required, policy)
                filled_path = _fill(template, record, idx, workdir)
                final_path = _maybe_convert(filled_path, run.output_format, workdir)
                output = _persist_output(run, final_path, idx, filename_field, record)
                outputs.append(output)
                _log(run, f"[OK]   ({idx + 1}/{len(records)}) wrote {output.filename}")
            except _RecordError as exc:
                run.report["errors"].append({"record": idx, "message": str(exc)})
                _log(run, f"[ERROR] record {idx}: {exc}")
                if policy == MissingPolicy.ERROR:
                    # Strict mode: one bad record fails the whole run.
                    raise
            except Exception as exc:  # unexpected per-record failure
                run.report["errors"].append({"record": idx, "message": repr(exc)})
                _log(run, f"[ERROR] record {idx} (unexpected): {exc}")

        # Batch rollup: one zip alongside the individual files.
        if len(outputs) > 1:
            _persist_archive(run, outputs)

    run.save(update_fields=["report"])
    return outputs


def _validate_record(run, idx, record, required, policy) -> None:
    """Compare a record against the template's placeholders; record gaps."""
    keys = set(record.keys())
    missing = sorted(required - keys)
    extra = sorted(keys - required)
    if extra:
        run.report["extra"][str(idx)] = extra
    if missing:
        run.report["missing"][str(idx)] = missing
        if policy == MissingPolicy.ERROR:
            raise _RecordError(f"missing placeholders: {missing}")
        # BLANK policy: fill gaps so the template renders cleanly.
        for k in missing:
            record[k] = ""


def _fill(template: Template, context: dict, idx: int, workdir: str) -> str:
    """Render the template with one record. Returns the filled office-file path."""
    out_name = f"filled_{idx:05d}.{_INTERMEDIATE[template.fmt]}"
    out_path = os.path.join(workdir, out_name)

    if template.fmt == TemplateFormat.DOCX:
        _fill_docx(template.file.path, context, out_path)
    elif template.fmt == TemplateFormat.XLSX:
        _fill_xlsx(template.file.path, context, out_path)
    else:  # pragma: no cover - guarded by choices
        raise _RecordError(f"unsupported template format: {template.fmt}")
    return out_path


def _fill_docx(template_path: str, context: dict, out_path: str) -> None:
    """Render a .docx via docxtpl. Sandboxed Jinja env as defense-in-depth (spec §9)."""
    from docxtpl import DocxTemplate
    from jinja2.sandbox import SandboxedEnvironment

    doc = DocxTemplate(template_path)
    jinja_env = SandboxedEnvironment(autoescape=False)
    doc.render(context, jinja_env)
    doc.save(out_path)


def _fill_xlsx(template_path: str, context: dict, out_path: str) -> None:
    """Replace {{key}} substrings in string cells; leave formulas/numbers intact."""
    from openpyxl import load_workbook

    placeholder = re.compile(r"\{\{\s*(\w+)\s*\}\}")
    wb = load_workbook(template_path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{" in cell.value:
                    cell.value = placeholder.sub(
                        lambda m: str(context.get(m.group(1), m.group(0))), cell.value
                    )
    wb.save(out_path)


def _maybe_convert(src_path: str, fmt: str, workdir: str) -> str:
    """If PDF is requested, convert via LibreOffice headless; else pass through.
    Cold-start per conversion is accepted (spec §7) — no warm soffice process."""
    if fmt != OutputFormat.PDF:
        return src_path

    # LibreOffice shares a profile across concurrent runs, so give each call its
    # own throwaway profile dir to avoid collisions on a busy worker.
    profile = tempfile.mkdtemp(prefix="lo_profile_", dir=workdir)
    proc = subprocess.run(
        [
            "soffice", "--headless",
            f"-env:UserInstallation=file://{profile}",
            "--convert-to", "pdf", "--outdir", workdir, src_path,
        ],
        capture_output=True, timeout=120,
    )
    if proc.returncode != 0:
        raise _RecordError(f"LibreOffice PDF conversion failed: "
                           f"{proc.stderr.decode(errors='replace')[:300]}")
    pdf_path = os.path.join(workdir, os.path.splitext(os.path.basename(src_path))[0] + ".pdf")
    if not os.path.exists(pdf_path):
        raise _RecordError("LibreOffice reported success but no PDF was produced")
    return pdf_path


# ──────────────────────────────────────────────
# Persistence
# ──────────────────────────────────────────────
def _persist_output(run, path, idx, filename_field, record) -> Output:
    """Read the produced file off disk and save it as an Output row."""
    fmt = run.output_format
    base = _output_basename(idx, filename_field, record)
    filename = f"{base}.{fmt}"

    with open(path, "rb") as fh:
        data = fh.read()

    output = Output(
        run=run, fmt=fmt, filename=filename,
        size_bytes=len(data),
        record_index=idx if run.record_count > 1 else None,
    )
    output.file.save(filename, ContentFile(data), save=False)
    output.save()
    return output


def _persist_archive(run, outputs: list[Output]) -> Output:
    """Zip a batch's individual files into one downloadable archive Output."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for out in outputs:
            with out.file.open("rb") as fh:
                zf.writestr(out.filename, fh.read())
    buf.seek(0)
    archive = Output(
        run=run, fmt=run.output_format, is_archive=True,
        filename=f"{_FILENAME_SANITIZE.sub('_', run.template.name)}_{run.id}.zip",
        size_bytes=buf.getbuffer().nbytes,
    )
    archive.file.save(archive.filename, ContentFile(buf.read()), save=False)
    archive.save()
    _log(run, f"[OK]   bundled {len(outputs)} files into {archive.filename}")
    return archive


def _output_basename(idx, filename_field, record) -> str:
    """Pick a filesystem-safe base name: the configured field, else doc_NNNN."""
    raw = ""
    if filename_field:
        raw = str(record.get(filename_field, "")).strip()
    base = _FILENAME_SANITIZE.sub("_", raw).strip("_") if raw else ""
    return base or f"doc_{idx + 1:04d}"


# ──────────────────────────────────────────────
# Delivery hand-off
# ──────────────────────────────────────────────
def _enqueue_deliveries(run, outputs: list[Output]) -> None:
    """Create PENDING Delivery rows per the binding's config and enqueue them."""
    binding = run.binding
    if not binding or not binding.delivery_kind:
        return

    # Deliver the archive for batches, the single file otherwise.
    targets = [o for o in run.outputs.all() if o.is_archive] or outputs
    for output in targets:
        delivery = Delivery.objects.create(
            output=output,
            kind=binding.delivery_kind,
            target=_resolve_target(binding, run),
            status=DeliveryStatus.PENDING,
        )
        async_task("docrunner.deliveries.send_delivery", str(delivery.id))
        _log(run, f"[INFO] queued {delivery.kind} delivery for {output.filename}")


def _resolve_target(binding, run) -> dict:
    """Build the concrete delivery target from the binding's delivery_config.
    Kept minimal here; the deliveries module does credential resolution."""
    return dict(binding.delivery_config)


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────
class _RecordError(Exception):
    """A per-record failure (validation or fill/convert) — handled per missing_policy."""


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _finalize(run, status, started_monotonic) -> None:
    run.status = status
    run.finished_at = _now()
    run.duration_ms = int((time.monotonic() - started_monotonic) * 1000)
    run.save(update_fields=["status", "finished_at", "duration_ms"])


def _log(run, line: str) -> None:
    """Append a prefixed log line to the run's report (PyRunner-style stdout capture)."""
    run.report.setdefault("log", []).append(line)
    print(line)  # also captured in the worker's own stdout
